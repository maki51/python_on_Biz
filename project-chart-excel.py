
# EXCEL用ライブラリの読み込み
from calendar import c
import openpyxl as excel
from openpyxl import utils

import os.path

#from openpyxl.cell.cell import VALID_TYPES

#import pandas as pd

# 正規表現のライブラリ
import re

# 日付時間のライブラリ
import datetime
#from datetime import timedelta

MAX_BLANK_ROWS=50

EXCEL_PARM_SHEET_NAME='Params'
EXCEL_PARM_SHEET_HIDUKE_PARM_NAME='日付開始セル'
EXCEL_PARM_SHEET_HIDUKE_ROW_PARM_NAME='日付開始セルrow'
EXCEL_PARM_SHEET_HIDUKE_COLM_PARM_NAME='日付開始セルcolmn'
EXCEL_PARM_SHEET_TASK_KAISHIGYO_PARM_NAME='タスク項目開始行'
EXCEL_PARM_SHEET_TASK_NUMBER_COLM_PARM_NAME='タスク番号カラム'
#EXCEL_PARM_TASKNO_COLUMN=1

# 読み込みファイル
INFILE_EXCEL_FULLPATH='C:\\Users\\maki51\\Desktop\\スケジュール.xlsx'
INFILE_EXCEL_DEFFILE=""



def myIsHoriday(argT):
    if (type(argT)=='str'):
      argT= datetime.datetime.strptime(argT, '%Y-%m-%d %h:%M:%S')

    try:
        weekday = argT.weekday()    
    except:
        return False
    
    
    if (weekday==5):   # saturday
        retDate = True
    elif (weekday==6):   #sunday
        retDate = True
    else:
        retDate = False

    return retDate

def getMyParmsOnExcel(parmName):
    try:
        sheet = wb[EXCEL_PARM_SHEET_NAME]
    except KeyError:
        print (f'シート {EXCEL_PARM_SHEET_NAME} がありません')

    for i in range(1,10):
        cell=sheet.cell(row=i,column=1)
        #print (cell.value)
        
        if (cell.value==parmName):
            parmVal = sheet.cell(row=i,column=2).value
            return parmVal

    print (f'シート 「{EXCEL_PARM_SHEET_NAME}」 に 項目「{parmName}」がありません')
    return None

print ("******* Script starting ********")


wb = excel.load_workbook(INFILE_EXCEL_FULLPATH,data_only=True)

# *****************************************************************
# 日付の色
sheet = wb.worksheets[0]

c1 = getMyParmsOnExcel(EXCEL_PARM_SHEET_HIDUKE_COLM_PARM_NAME)
r1 = getMyParmsOnExcel(EXCEL_PARM_SHEET_HIDUKE_ROW_PARM_NAME)

ccc=0

daysOnProject={}

while (True):
    d = sheet.cell(row=r1,column=ccc+c1) 

    if (d.number_format=="General"):
        d = utils.datetime.from_excel(d.value)
        #https://texalog.com/archives/69
    else:
        d=d.value

    if (d==None):
        break

    #if (myIsHoriday(d) ):
    #    sheet.cell(row=r1,column=ccc).fill = excel.styles.PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
    #    sheet.cell(row=r1,column=ccc).font = excel.styles.fonts.Font(color='FFFFFF')
    #else :
    #    sheet.cell(row=r1,column=ccc).fill = excel.styles.PatternFill(patternType='solid',fgColor='FFFFFF', bgColor='FFFFFF')
    #    sheet.cell(row=r1,column=ccc).font = excel.styles.fonts.Font(color='000000')

    daysOnProject[ccc]=d
    #print(ccc,d)
    ccc+=1
    

# プロジェクト期間の日数を取得
projectPeriod = ccc
print (f'ProjectPeriod {projectPeriod}days')



# *****************************************************************
# タスク番号の取得
sheet = wb.worksheets[0]

c1 = getMyParmsOnExcel(EXCEL_PARM_SHEET_TASK_NUMBER_COLM_PARM_NAME)
r1 = getMyParmsOnExcel(EXCEL_PARM_SHEET_TASK_KAISHIGYO_PARM_NAME)

rrr=r1

taskRowNumber_dict={}
noneFlag = 0

while (True):
    d = sheet.cell(row=rrr,column=c1) 
    
    d=d.value

    #空白行が2行以上連続して続いていたら ループから抜ける
    if (d==None):
        noneFlag +=1
        if (noneFlag >=2):
            break
    else:
        noneFlag=0
        taskRowNumber_dict[d] = rrr
    
    rrr+=1

#print (taskRowNumber_dict)


# *****************************************************************
# タスクの開始日の確認
sheet = wb.worksheets[0]

c1 = getMyParmsOnExcel(EXCEL_PARM_SHEET_HIDUKE_COLM_PARM_NAME)
r1 = getMyParmsOnExcel(EXCEL_PARM_SHEET_TASK_KAISHIGYO_PARM_NAME)


ccc=c1
rrr=r1

#while (True):
#    d = sheet.cell(row=r1,column=ccc) 


taskMinDay_dict={}
taskMaxDay_dict={}
taskPreTaskNumbers={}
#taskPreTaskNumbers=list(range(0))

for taskNo in taskRowNumber_dict:
     
    maxDay=0-1
    minDay=getMyParmsOnExcel(EXCEL_PARM_SHEET_TASK_NUMBER_COLM_PARM_NAME)+projectPeriod+1
    #preTask={}
    preTask =[]


    for ddd in range(1,projectPeriod):
        d = sheet.cell(row=taskRowNumber_dict[taskNo],column=c1+ddd) 

        d=d.value
        if (d==None): 
            continue
        else: 
            d=str(d)

        mtc = re.match('^[a-zA-Z]+',d )

        if (mtc):
            if (minDay >ddd):
                minDay=ddd
            
            if(maxDay<ddd):
                maxDay = ddd


        mtc = re.match('^[0-9]+',d )

        #　先行タスク番号の取得
        if(mtc):
            #print (d)
            preTask.append(d)


    taskMinDay_dict[taskNo]=minDay
    taskMaxDay_dict[taskNo]=maxDay

    taskPreTaskNumbers[taskNo]=preTask

    #print (f'TaskNo:{taskNo}\trow:{taskRowNumber_dict[taskNo]} \tMin:{taskMinDay_dict[taskNo]} \tMax: {taskMaxDay_dict[taskNo]}')
    #print (taskPreTaskNumbers[taskNo])

    #print (f'Min {minDay}  Max{maxDay}')    
    #print ('------------------------')



#  先行タスクの終了日よりも タスクの開始日が後になっているかのチェック
for taskNo in taskRowNumber_dict:
  for pret in taskPreTaskNumbers[taskNo]:
      pret = int(pret)
      #print (taskMinDay_dict[pret])
      if (taskMinDay_dict[taskNo] <  taskMaxDay_dict[pret]) :
        print ('\n------------------------------------------------------')
        print (f'TaskNo.{taskNo} must be started later than {daysOnProject[taskMinDay_dict[taskNo]]},') 
        print (f'cause Pre-Task Number.{pret} is finished on {daysOnProject[taskMaxDay_dict[pret]]}.' )
        
  
print ()




file1=os.path.basename(INFILE_EXCEL_FULLPATH)
path1=os.path.dirname(INFILE_EXCEL_FULLPATH)
file1 = '\\reported_' + file1

fullpath1= os.path.join(path1 + file1)
#print(fullpath1)

#wb.save(fullpath1)





